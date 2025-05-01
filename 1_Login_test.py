from tkinter import *
from tkinter import PhotoImage
from subprocess import call
from tkinter import messagebox
from openpyxl import load_workbook

window = Tk()
window.title("Login Portal")
window.geometry('1300x825')
window.resizable(False, False)

# ==================== Functions ====================

# Login Checker Function
def login_checker():
    if not data_validation_debugger():
        return

    username = user_entry.get()
    password = password_entry.get()

    try:
        wb = load_workbook("user_account_data.xlsx")
        ws = wb["Userdata"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username and row[1] == password:
                messagebox.showinfo("Login Success", f"Welcome, {username}!")
                window.destroy()  #Close the login window
                call(["python", "3_Enrollment_Form_test.py"])  # Open the next page/code
                return

    
        messagebox.showerror(title= "Login Failed", message= "Incorrect username or password.")

    except FileNotFoundError:
        messagebox.showerror(title= "Error", message= "User data file not found.")

# Input Validation & Debugger Function
def data_validation_debugger():

    username = user_entry.get()
    password = password_entry.get()

    if username and password and username != "Username" and password != "Password":
        print("\n\nData Entry Form:\n\n" \
              "Username:",username,
              "\nPassword:",password,
              "\n")
        return True
    
    else:
        messagebox.showerror(title= "Error", message= "Error. Input Required")
        return False

# Username Delete & Restore Function
def on_username_click(event):
    if user_entry.get() == "Username":
        user_entry.delete(0, END)

def on_username_leave(event):
    name = user_entry.get()
    if name == "":
        user_entry.insert(0, "Username")

# Password Delete & Restore Function
def on_password_click(event):
    if password_entry.get() == "Password":
        password_entry.delete(0, END)

def on_password_leave(event):
    password = password_entry.get()
    if password == "":
        password_entry.insert(0, "Password")

# Signup Window Switch Function
def go_signup():
    window.destroy()
    call(["python", "2_Sign_Up_test.py"])

def handle_enter(event):
    if not login_checker():
        return

# //////////////////////////////////////////////////////////

# ==================== UI ====================
# Background & Banner Img
image_path = PhotoImage(file= r'.\wallhaven-85gxp2.png')
bg_image = Label(window, image= image_path)
bg_image.place(relheight=1, relwidth=1)

image_file = PhotoImage(file = r'.\wallhaven-73616y.png')
image = image_file.subsample(6, 4)

Label(window, image= image, highlightbackground="black", highlightthickness=4).grid(row=0, column=0, padx=70, pady=140)

# //////////////////////////////////////////////////////////

# ==================== Frames ====================
frame = Frame(window, bg= "white", highlightbackground="black", highlightthickness=4)
frame.grid(row=0, column=1, padx=45)

# //////////////////////////////////////////////////////////

# ==================== Labels & Inputs ====================

# Project Name
project_label = Label(frame, text= "PROJECT", font= ("Times New Roman bold", 30), bg= "white")
project_label.grid(row=0, column=0, sticky="n", pady= 10)

# Short Description
short_desc_label =  Label(frame, text= "/Short Description/", font= ("Helvetica bold", 15), bg= "white")
short_desc_label.grid(row=1, column=0, sticky="n")

# Login
login_label = Label(frame, text= "Log In to Project", font= ("Microsoft YaHei UI Light", 18, "bold"), bg= "white")
login_label.grid(row=2, column=0, sticky="w", pady=15, padx= 15)

# User Entry
user_entry = Entry(frame, width=38, font= ("Arial", 15), bd=0)
user_entry.grid(row=3, column=0, pady=1, sticky="n")
user_entry.insert(0, "Username")
user_entry.bind("<FocusIn>", on_username_click)
user_entry.bind("<FocusOut>", on_username_leave)
user_entry.bind('<Return>', handle_enter)

# Underline using Frame
underline = Frame(frame, width=464, height=2, bg="black")
underline.place(x=17, y=198)

# Password Entry
password_entry = Entry(frame, width=38, font= ("Arial", 15), bd=0)
password_entry.grid(row=4, column=0, pady=37, sticky="n")
password_entry.insert(0, "Password")
password_entry.bind("<FocusIn>", on_password_click)
password_entry.bind("<FocusOut>", on_password_leave)
password_entry.bind('<Return>', handle_enter)

# Underline using Frame
underline = Frame(frame, width=464, height=2, bg="black")
underline.place(x=17, y=262)

# //////////////////////////////////////////////////////////

# ==================== Buttons ====================

# Button Login
login_button = Button(frame, text= "Log In", width=38, font= ("Arial bold", 15), bg= "dodger blue", fg= "white", command= login_checker)
login_button.grid(row=5, column=0, pady=5, padx= 15)

# SIGNUP Button
need_account = Label(frame, text="Need an Account?", font=("Arial", 12), bg= "white")
need_account.grid(row=6, column=0, sticky="nw", pady=15, padx=145)

signup_button = Button(frame, text="SIGN UP", font=("Arial", 12), fg= "dodgerblue2", bg= "white", borderwidth= 0, command=go_signup)
signup_button.grid(row=6, column=0, sticky="ne", pady=15, padx=145)


# //////////////////////////////////////////////////////////


frame.columnconfigure(0, weight=1)
window.configure(bg="white")
window.mainloop()