from tkinter import *
from tkinter import PhotoImage
from subprocess import call
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

window = Tk()
window.title("Login Portal")
window.geometry('1300x825')
window.resizable(False, False)


# ==================== Functions ====================

# Create and Save to Excel Function
def save_to_excel():
    if not data_validation_debugger():
        return
    
    username = user_entry.get()
    password = password_entry.get()

    try:
        wb = load_workbook("user_account_data.xlsx")
        if "Userdata" in wb.sheetnames:
            ws = wb["Userdata"]
        else:
            ws = wb.create_sheet("Userdata")

    except FileNotFoundError:

        wb = Workbook()
        ws = wb.active
        ws.title = "Userdata"
        ws.append(["Username", "Password"])

    ws.append([username, password])
    wb.save("user_account_data.xlsx")

    format_excel()
    show_data()

    messagebox.showinfo(title= "Success", message= "Account Saved")
    user_entry.delete(0, END)
    password_entry.delete(0, END)
    user_entry.insert(0, "Username")
    password_entry.insert(0, "Password")
     
# Format Fixer Function
def format_excel():
    wb = load_workbook("user_account_data.xlsx")
    ws = wb["Userdata"]

    # Bold Headers
    for cells in ws[1]:
        cells.font = Font(bold=True)

    # Auto Column Width
    for cols in ws.columns:
        max_length = max(len(str(cells.value)) for cell in cols)
        col_letter = get_column_letter(cols[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save("user_account_data.xlsx")

# New Window Data Shower Function
def show_data():
    wb = load_workbook("user_account_data.xlsx")
    ws = wb["Userdata"]

    data_window = Toplevel(window)
    data_window.title("Stored User Data")

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        for j, value in enumerate(row):
            label = Label(data_window, text=value, borderwidth=1, relief="solid", padx=6, pady=3)
            label.grid(row=i, column=j)

# Input Validation & Debugger Function
def data_validation_debugger():
    username = user_entry.get()
    password = password_entry.get()

    if username and password and username != "Username" and password != "Password":
        print("\n\nSignup Form:\n\nUsername:", username, "\nPassword:", password, "\n")
        return True
    
    else:
        messagebox.showerror(title= "Error", message= "Please enter both Username and Password.")
        return False

# Username Delete & Restore Function
def on_username_click(event):
    if user_entry.get() == "Username":
        user_entry.delete(0, END)

def on_username_leave(event):
    name = user_entry.get()
    if name == "":
        user_entry.insert(0, "Username")

# Password Delete & Restore Function
def on_password_click(event):
    if password_entry.get() == "Password":
        password_entry.delete(0, END)

def on_password_leave(event):
    password = password_entry.get()
    if password == "":
        password_entry.insert(0, "Password")

# Signup Window Switch Function
def go_login():
    window.destroy()
    call(["python", "1_Login_test.py"])

def handle_enter(event):
    if not save_to_excel():
        return

# //////////////////////////////////////////////////////////

# ==================== UI ====================
# Background & Banner Img
image_path = PhotoImage(file= r'.\wallhaven-85gxp2.png')
bg_image = Label(window, image= image_path)
bg_image.place(relheight=1, relwidth=1)

# //////////////////////////////////////////////////////////

# ==================== Frames ====================
frame = Frame(window, bg= "white", highlightbackground="black", highlightthickness=4)
frame.place(relx = 0.5, rely = 0.5, anchor = CENTER)

# //////////////////////////////////////////////////////////

# ==================== Labels & Inputs ====================

# Project Name
project_label = Label(frame, text= "PROJECT", font= ("Times New Roman bold", 30), bg= "white")
project_label.grid(row=0, column=0, sticky="n", pady= 10)

# Short Description
short_desc_label =  Label(frame, text= "/Short Description/", font= ("Helvetica bold", 15), bg= "white")
short_desc_label.grid(row=1, column=0, sticky="n")

# Sign Up
signup_label = Label(frame, text= "Sign Up", font= ("Microsoft YaHei UI Light", 18, "bold"), bg= "white")
signup_label.grid(row=2, column=0, sticky="w", pady=15, padx= 15)

# User Entry
user_entry = Entry(frame, width=38, font= ("Arial", 15), bd=0)
user_entry.grid(row=3, column=0, pady=1, sticky="n")
user_entry.insert(0, "Username")
user_entry.bind("<FocusIn>", on_username_click)
user_entry.bind("<FocusOut>", on_username_leave)
user_entry.bind('<Return>', handle_enter)

# Underline using Frame
underline = Frame(frame, width=464, height=2, bg="black")
underline.place(x=17, y=198)

# Password Entry
password_entry = Entry(frame, width=38, font= ("Arial", 15), bd=0)
password_entry.grid(row=4, column=0, pady=37, sticky="n")
password_entry.insert(0, "Password")
password_entry.bind("<FocusIn>", on_password_click)
password_entry.bind("<FocusOut>", on_password_leave)
password_entry.bind('<Return>', handle_enter)

# Underline using Frame
underline = Frame(frame, width=464, height=2, bg="black")
underline.place(x=17, y=262)

# //////////////////////////////////////////////////////////

# ==================== Buttons ====================

# Button Sign Up
signin_button = Button(frame, text= "Sign Up", width=38, font= ("Arial bold", 15), bg= "dodger blue", fg= "white", command= save_to_excel)
signin_button.grid(row=5, column=0, pady=5, padx= 15)

# LOGIN Button
already_user = Label(frame, text="Already a User?", font=("Arial", 12), bg= "white")
already_user.grid(row=6, column=0, sticky="nw", pady=15, padx=161)

login_button = Button(frame, text="LOGIN", font=("Arial", 12), fg= "dodgerblue2", bg= "white", borderwidth= 0, command= go_login)
login_button.grid(row=6, column=0, sticky="ne", pady=13, padx=161)



window.mainloop()